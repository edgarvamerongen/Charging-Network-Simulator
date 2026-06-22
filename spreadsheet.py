"""
CNS — XLSX spreadsheet export of the Demand Calculator.

A standardised, responsive workbook that mirrors the PDF report's data and
figures. Deliberately SEPARATE from report.py: it only *reads* the same
self-contained payload the browser already assembles (plus a richer
`flightsFull` list), and never touches the PDF pipeline or the JS engine.

Workbook v2 layout — three kinds of sheets:
  * `Overview` — KPI cards, a hyperlinked airport index, and brand-styled
    horizontal bar charts (daily energy / peak per airport).
  * `Data` — the About/version block plus ALL five tables stacked:
    Airports (computed) and the round-trippable inputs Flights / Aircraft /
    Chargers / Settings. A FUTURE importer reads the input tables — they are
    addressed by Excel Table NAME (tblFlights, …), so their sheet placement is
    irrelevant to the contract. `CNS_Version` is a defined name an importer
    validates first.
  * one tab per airport — KPIs, installed charging, contributing flights,
    energy-mix donut + step-true load curve, charts anchored CLEAR of the
    tables (column K onward).

Responsiveness: daily energy, revenue, installed power, totals and % mix are
native Excel formulas (recompute when the user edits a frequency, tariff or
charger); charts are bound to cell ranges. Scheduler-derived figures
(per-flight charge energy, peak kW, the load-curve series) can't be expressed
as cell formulas, so they're exported as values — charts over them still
redraw if the series is edited.

v2 vs v1 (breaking → version bump): the five category sheets merged into
`Data`; `tblSettings` changed from key/value rows to a SINGLE-ROW table (one
column per setting); About sheet folded into `Data`.
"""
import io
from datetime import datetime

from economics import (DAY_START_MIN, DAY_END_MIN, REALISATION_LOW,
                       REALISATION_HIGH, PROCUREMENT_EUR_PER_KWH)

FORMAT_NAME = 'NRG2FLY Charging Network Simulator — workbook'
FORMAT_VERSION = 'CNS Workbook v2'

# ---- styling (Arial per xlsx conventions; house palette for accents) --------
FONT = 'Arial'
C_INPUT = 'FF0000FF'    # blue text — hardcoded inputs the user may change
C_FORMULA = 'FF000000'  # black text — formulas / calculations
C_LINK = 'FF008000'     # green text — links pulling from another sheet
C_MUTED = 'FF6B7280'
NAVY = 'FF152455'
BLUE = 'FF2563EB'
ORANGE = 'FFF0892B'
GREEN = 'FF10B981'
SOFT = 'FFF1F5F9'
WHITE = 'FFFFFFFF'
TAB_NAVY = '152455'
TAB_GREEN = '217346'    # classic spreadsheet green (matches the export button)
TAB_AIRPORT = 'BFD2F8'  # light blue family for the airport tab group

# donut slice palette (hex without alpha, as chart fills want them)
PALETTE = ['2563EB', 'F0892B', '10B981', '6F42C1', '0EA5E9',
           'F59E0B', '14B8A6', 'EF4444', '8B5CF6', '64748B']

FMT_KWH = '#,##0 "kWh"'
FMT_KW = '#,##0 "kW"'
FMT_EUR = '€#,##0'
FMT_PCT = '0.0%'
FMT_COORD = '0.0000'
FMT_NUM = '#,##0'
FMT_TIME = 'hh:mm'


def _fpd_expr(n_ref, unit_ref):
    """Flights/day as a live formula: weekly trips amortise over 7 days."""
    return f'=IF({unit_ref}="week",{n_ref}/7,{n_ref})'


class SpreadsheetBuilder:
    def __init__(self, payload):
        from openpyxl import Workbook
        self.p = payload or {}
        self.wb = Workbook()
        # busiest-first everywhere: tab order, index, charts all read nicer
        self.airports = sorted(self.p.get('airports') or [],
                               key=lambda a: float(a.get('dailyKwh') or 0), reverse=True)
        self.planes = self.p.get('planes') or []
        self.chargers = self.p.get('chargers') or []
        self.flights = self.p.get('flightsFull') or self.p.get('flights') or []
        self.settings = self.p.get('modelSettings') or {}
        try:
            self.tariff = float(self.p.get('chargeRate'))
        except (TypeError, ValueError):
            self.tariff = 0.60
        self.realisation_low = REALISATION_LOW
        self.realisation_high = REALISATION_HIGH
        self.procurement = PROCUREMENT_EUR_PER_KWH
        self._tab_names = set()
        # filled while building per-airport sheets; consumed by Data + Overview
        self._airport_refs = []   # [{ident, name, lat, lon, sheet, daily, peak, installed}]
        self._ap_rows = []        # Data-sheet row number per airport (index order)

    # ---- low-level helpers --------------------------------------------------
    def _cell(self, ws, r, c, value=None, *, bold=False, size=11, color=C_FORMULA,
              fill=None, fmt=None, align=None, wrap=False, italic=False):
        from openpyxl.styles import Font, PatternFill, Alignment
        cell = ws.cell(row=r, column=c)
        if value is not None:
            cell.value = value
        cell.font = Font(name=FONT, bold=bold, size=size, color=color, italic=italic)
        if fill:
            cell.fill = PatternFill('solid', fgColor=fill)
        if fmt:
            cell.number_format = fmt
        if align or wrap:
            cell.alignment = Alignment(horizontal=align or 'left', vertical='center', wrap_text=wrap)
        return cell

    def _header_row(self, ws, r, labels, start_col=1):
        for i, lab in enumerate(labels):
            self._cell(ws, r, start_col + i, lab, bold=True, color=WHITE, fill=NAVY, align='left')

    def _section(self, ws, r, label, tag):
        """A section band on the Data sheet: navy label + muted input/computed tag."""
        self._cell(ws, r, 1, label, bold=True, size=12, color=NAVY)
        self._cell(ws, r, 2, tag, italic=True, size=9, color=C_MUTED)

    def _table(self, ws, name, first_row, last_row, first_col, last_col):
        """Wrap a written range in an Excel Table (import anchor + styling)."""
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        ref = f'{get_column_letter(first_col)}{first_row}:{get_column_letter(last_col)}{last_row}'
        tbl = Table(displayName=name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(name='TableStyleLight9', showRowStripes=True,
                                            showFirstColumn=False, showLastColumn=False)
        ws.add_table(tbl)

    def _named(self, name, ws_title, cell_ref):
        from openpyxl.workbook.defined_name import DefinedName
        self.wb.defined_names.add(DefinedName(name, attr_text=f"'{ws_title}'!{cell_ref}"))

    def _widths(self, ws, widths):
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _tab(self, raw):
        """A valid, unique worksheet tab name (≤31 chars, no []:*?/\\)."""
        base = ''.join(ch for ch in str(raw or 'Airport') if ch not in '[]:*?/\\').strip()[:31] or 'Airport'
        name, i = base, 2
        while name.lower() in self._tab_names:
            suffix = f' ({i})'
            name = base[:31 - len(suffix)] + suffix
            i += 1
        self._tab_names.add(name.lower())
        return name

    def _hyperlink(self, ws, r, c, text, target_sheet):
        """An internal hyperlink cell that jumps to another sheet."""
        from openpyxl.worksheet.hyperlink import Hyperlink
        cell = self._cell(ws, r, c, _safe_text(text), color=BLUE)
        from openpyxl.styles import Font
        cell.font = Font(name=FONT, color=BLUE, underline='single')
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{target_sheet}'!A1")
        return cell

    # ---- chart styling helpers ----------------------------------------------
    @staticmethod
    def _solid_series(series, hex_color, line=False, width_emu=28575):
        """One brand colour for a whole series (kills openpyxl's rainbow default)."""
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties
        if line:
            gp = GraphicalProperties()
            gp.line = LineProperties(solidFill=hex_color, w=width_emu)
            series.graphicalProperties = gp
        else:
            series.graphicalProperties = GraphicalProperties(solidFill=hex_color)

    # ---- per-airport sheets --------------------------------------------------
    def _airport_detail(self, a):
        name = a.get('name') or a.get('ident') or 'Airport'
        ident = a.get('ident') or ''
        ws = self.wb.create_sheet(self._tab(ident or name))
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = TAB_AIRPORT
        self._widths(ws, [30, 14, 16, 16, 14, 12, 12, 14, 14, 4])
        self._cell(ws, 1, 1, _safe_text(name), bold=True, size=16, color=NAVY)
        self._cell(ws, 2, 1, _safe_text(ident), bold=True, color=C_MUTED)
        if a.get('lat') is not None:
            self._cell(ws, 2, 2, f"{_num(a.get('lat')):.4f}, {_num(a.get('lon')):.4f}", color=C_MUTED)

        contribs = a.get('contribs') or []
        chargers = a.get('chargers') or []

        # --- Installed charging ---
        r = 4
        self._cell(ws, r, 1, 'INSTALLED CHARGING', bold=True, color=NAVY)
        r += 1
        self._header_row(ws, r, ['Charger', 'Count', 'Power each (kW)', 'Total (kW)'])
        ch_first = r + 1
        r += 1
        for c in chargers:
            self._cell(ws, r, 1, _safe_text(c.get('name')), color=C_INPUT)
            self._cell(ws, r, 2, _num(c.get('count')) or 1, color=C_INPUT, fmt=FMT_NUM)
            self._cell(ws, r, 3, _num(c.get('power_kw')), color=C_INPUT, fmt=FMT_NUM)
            self._cell(ws, r, 4, f'=B{r}*C{r}', color=C_FORMULA, fmt=FMT_KW)
            r += 1
        ch_last = r - 1
        installed_cell = f'D{r}'
        self._cell(ws, r, 1, 'Total installed power', bold=True, color=NAVY)
        self._cell(ws, r, 4, (f'=SUM(D{ch_first}:D{ch_last})' if ch_last >= ch_first else 0),
                   bold=True, color=C_FORMULA, fmt=FMT_KW)
        r += 2

        # --- Contributing flights (energy/flight = value; daily = formula) ---
        self._cell(ws, r, 1, 'CONTRIBUTING FLIGHTS', bold=True, color=NAVY)
        r += 1
        self._header_row(ws, r, ['Route', 'Aircraft', 'Trip', 'Freq N', 'Freq unit',
                                 'Energy/flight (kWh)', 'Flights/day', 'Daily total (kWh)', 'Charge time (min)'])
        cf_first = r + 1
        r += 1
        ac_col, daily_col = 2, 8
        for c in contribs:
            role = (c.get('role') or '').upper()
            other = c.get('other') or ''
            route = {'HOME': f'{role} → {other} & back', 'ORIGIN': f'{role} → {other}',
                     'STOP': f'{role} on {other}'}.get(role, f'{role} from {other}')
            trip = ('Return' if c.get('tripType') == 'retour' else 'One-way') + (' · multi-leg' if c.get('multiLeg') else '')
            self._cell(ws, r, 1, _safe_text(route))
            self._cell(ws, r, 2, _safe_text(c.get('planeName')))
            self._cell(ws, r, 3, trip)
            self._cell(ws, r, 4, _num(c.get('freqN')), color=C_INPUT, fmt=FMT_NUM)
            self._cell(ws, r, 5, _safe_text(c.get('freqUnit')), color=C_INPUT)
            self._cell(ws, r, 6, _num(c.get('energyPerFlight')), color=C_INPUT, fmt=FMT_NUM)
            self._cell(ws, r, 7, _fpd_expr(f'D{r}', f'E{r}'), color=C_FORMULA, fmt='0.00')
            self._cell(ws, r, 8, f'=F{r}*G{r}', color=C_FORMULA, fmt=FMT_NUM)
            self._cell(ws, r, 9, _num(c.get('chargeMin')), color=C_INPUT, fmt=FMT_NUM)
            r += 1
        cf_last = r - 1
        daily_total_cell = f'H{r}'
        self._cell(ws, r, 1, 'Daily total', bold=True, color=NAVY)
        self._cell(ws, r, daily_col, (f'=SUM(H{cf_first}:H{cf_last})' if cf_last >= cf_first else 0),
                   bold=True, color=C_FORMULA, fmt=FMT_KWH)
        r += 2

        # --- KPI header cells (top-right of the title block) ---
        self._cell(ws, 1, 3, 'Revenue/day', bold=True, color=C_MUTED, align='right')
        self._cell(ws, 1, 4, f'={daily_total_cell}*Settings_tariff', color=C_FORMULA, fmt=FMT_EUR, align='right')
        self._cell(ws, 2, 3, 'Peak', bold=True, color=C_MUTED, align='right')
        self._cell(ws, 2, 4, _num(a.get('peakKw')), color=C_INPUT, fmt=FMT_KW, align='right')
        self._cell(ws, 3, 3, 'Daily', bold=True, color=C_MUTED, align='right')
        self._cell(ws, 3, 4, f'={daily_total_cell}', color=C_LINK, fmt=FMT_KWH, align='right')

        # --- Energy by aircraft type (SUMIF over the contrib rows) ---
        types = []
        for c in contribs:
            t = c.get('planeName') or 'Unknown'
            if t not in types:
                types.append(t)
        self._cell(ws, r, 1, 'ENERGY BY AIRCRAFT TYPE', bold=True, color=NAVY)
        r += 1
        self._header_row(ws, r, ['Aircraft', 'Daily (kWh)', '% of day'])
        ebt_data_first = r + 1
        r += 1
        from openpyxl.utils import get_column_letter as _gcl
        ac_rng = f'${_gcl(ac_col)}${cf_first}:${_gcl(ac_col)}${cf_last}'
        daily_rng = f'${_gcl(daily_col)}${cf_first}:${_gcl(daily_col)}${cf_last}'
        for t in types:
            if cf_last >= cf_first:
                self._cell(ws, r, 1, _safe_text(t))
                self._cell(ws, r, 2, f'=SUMIF({ac_rng},A{r},{daily_rng})', color=C_FORMULA, fmt=FMT_NUM)
                self._cell(ws, r, 3, f'=IF({daily_total_cell}=0,0,B{r}/{daily_total_cell})', color=C_FORMULA, fmt=FMT_PCT)
                r += 1
        ebt_data_last = r - 1

        # --- Load curve: STEP series on an Excel-time axis -------------------
        # The payload's breakpoints say "power is kw from t until the next t".
        # A plain line chart draws diagonals between them, so we write doubled
        # points — (t, prev) then (t, new) — and chart them as a scatter with
        # straight lines: a true step on a time-PROPORTIONAL x-axis.
        curve = a.get('loadCurvePoints') or []
        lc_data_first = lc_data_last = None
        if curve:
            r += 1
            self._cell(ws, r, 1, 'DAILY LOAD PROFILE', bold=True, color=NAVY)
            self._cell(ws, r, 2, 'step series — power holds until the next time',
                       italic=True, size=9, color=C_MUTED)
            r += 1
            self._header_row(ws, r, ['Time', 'Power (kW)'])
            lc_data_first = r + 1
            r += 1
            prev = None
            for pt in curve:
                t = _num(pt.get('t'))
                kw = _num(pt.get('kw')) or 0
                if t is None:
                    continue
                if prev is not None and kw != prev:
                    self._cell(ws, r, 1, t / 1440.0, color=C_INPUT, fmt=FMT_TIME)
                    self._cell(ws, r, 2, prev, color=C_INPUT, fmt=FMT_NUM)
                    r += 1
                self._cell(ws, r, 1, t / 1440.0, color=C_INPUT, fmt=FMT_TIME)
                self._cell(ws, r, 2, kw, color=C_INPUT, fmt=FMT_NUM)
                prev = kw
                r += 1
            lc_data_last = r - 1

        # --- charts: anchored at column K, CLEAR of the A–I tables -----------
        self._add_donut(ws, ebt_data_first, ebt_data_last, anchor='K2')
        if lc_data_first and lc_data_last and lc_data_last > lc_data_first:
            self._add_step_curve(ws, lc_data_first, lc_data_last, anchor='K20')

        self._airport_refs.append({
            'ident': ident, 'name': name, 'lat': a.get('lat'), 'lon': a.get('lon'),
            'sheet': ws.title, 'daily': daily_total_cell, 'peak': _num(a.get('peakKw')),
            'installed': installed_cell,
        })
        return ws

    # ---- Data sheet (About block + all five tables) ---------------------------
    def _data_sheet(self):
        ws = self.wb.create_sheet('Data')
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = TAB_GREEN
        self._widths(ws, [24, 20, 14, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 12, 12])

        # --- About block ---
        self._cell(ws, 1, 1, 'NRG2FLY · Charging Network Plan', bold=True, size=16, color=NAVY)
        self._cell(ws, 2, 1, 'Format', bold=True, color=NAVY)
        self._cell(ws, 2, 2, FORMAT_NAME, color=C_MUTED)
        self._cell(ws, 3, 1, 'Version', bold=True, color=NAVY)
        self._cell(ws, 3, 2, FORMAT_VERSION, color=C_MUTED)
        self._named('CNS_Version', 'Data', '$B$3')
        self._cell(ws, 4, 1, 'Generated', bold=True, color=NAVY)
        self._cell(ws, 4, 2, _safe_text(self.p.get('generatedAt') or datetime.now().strftime('%Y-%m-%d %H:%M')),
                   color=C_MUTED)
        self._cell(ws, 5, 1, 'Scope', bold=True, color=NAVY)
        self._cell(ws, 5, 2, f'{len(self.airports)} airports · {len(self.flights)} flights', color=C_MUTED)
        note = ('Input tables (Flights, Aircraft, Chargers, Settings) are the standardised, round-trippable '
                'plan — a future import reads only these, found by table name. Computed content (Overview, '
                'the Airports table, airport tabs) is regenerated. Live figures (daily energy, revenue, '
                'installed power, % mix) are Excel formulas; per-flight charge energy, peak power and the '
                'load curves come from the CNS simulator and are exported as values — re-run the simulator '
                'to recompute those.')
        self._cell(ws, 6, 1, 'How to read', bold=True, color=NAVY)
        cell = self._cell(ws, 6, 2, note, size=9, color=C_MUTED, wrap=True)
        ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=10)
        ws.row_dimensions[6].height = 52

        r = 8

        # --- AIRPORTS (computed) ---
        self._section(ws, r, 'AIRPORTS', 'computed — regenerated on import')
        r += 1
        cols = ['Airport', 'ICAO', 'Lat', 'Lon', 'Daily energy (kWh)', 'Peak (kW)',
                'Installed (kW)', 'Revenue/day (EUR)']
        self._header_row(ws, r, cols)
        first = r + 1
        r += 1
        self._ap_rows = []
        for ref in self._airport_refs:
            q = f"'{ref['sheet']}'"
            self._cell(ws, r, 1, _safe_text(ref['name']))
            self._cell(ws, r, 2, _safe_text(ref['ident']))
            self._cell(ws, r, 3, _num(ref['lat']), fmt=FMT_COORD)
            self._cell(ws, r, 4, _num(ref['lon']), fmt=FMT_COORD)
            self._cell(ws, r, 5, f"={q}!{ref['daily']}", color=C_LINK, fmt=FMT_NUM)
            self._cell(ws, r, 6, ref['peak'], color=C_INPUT, fmt=FMT_NUM)
            self._cell(ws, r, 7, f"={q}!{ref['installed']}", color=C_LINK, fmt=FMT_NUM)
            self._cell(ws, r, 8, f'=E{r}*Settings_tariff', color=C_FORMULA, fmt=FMT_EUR)
            self._ap_rows.append(r)
            r += 1
        self._ap_first, self._ap_last = first, max(r - 1, first)
        self._table(ws, 'tblAirports', first - 1, self._ap_last, 1, len(cols))
        r += 2

        # --- FLIGHTS (input) ---
        self._section(ws, r, 'FLIGHTS', 'input — round-trippable')
        r += 1
        cols = ['Flight ID', 'Aircraft ID', 'Aircraft', 'Origin ICAO', 'Origin', 'Origin Lat',
                'Origin Lon', 'Dest ICAO', 'Dest', 'Dest Lat', 'Dest Lon', 'Stops (ICAO;…)',
                'Trip type', 'Multi-leg', 'Charger ID', 'Charger', 'Freq N', 'Freq unit']
        self._header_row(ws, r, cols)
        hdr = r
        r += 1
        for f in self.flights:
            stops = ';'.join(s.get('ident', '') for s in (f.get('stops') or []) if isinstance(s, dict))
            vals = [f.get('id'), f.get('planeId'), f.get('planeName'),
                    f.get('originIdent'), f.get('originName'), _num(f.get('originLat')), _num(f.get('originLon')),
                    f.get('destIdent'), f.get('destName'), _num(f.get('destLat')), _num(f.get('destLon')),
                    stops, f.get('tripType'), 'yes' if f.get('multiLeg') else 'no',
                    f.get('chargerId'), f.get('chargerName'), _num(f.get('freqN')), f.get('freqUnit')]
            for i, v in enumerate(vals, start=1):
                fmt = FMT_COORD if i in (6, 7, 10, 11) else None
                self._cell(ws, r, i, _safe_text(v), color=C_INPUT, fmt=fmt)
            r += 1
        self._table(ws, 'tblFlights', hdr, max(r - 1, hdr + 1), 1, len(cols))
        r += 2

        # --- AIRCRAFT (input) ---
        self._section(ws, r, 'AIRCRAFT', 'input — round-trippable')
        r += 1
        cols = ['Aircraft ID', 'Name', 'Battery (kWh)', 'Range (km)', 'Speed (km/h)', 'Seats', 'Payload (kg)']
        self._header_row(ws, r, cols)
        hdr = r
        r += 1
        for pl in self.planes:
            vals = [pl.get('id'), pl.get('name'), _num(pl.get('battery_kwh')), _num(pl.get('range_km')),
                    _num(pl.get('speed_kmh')), _num(pl.get('seats')), _num(pl.get('load_kg'))]
            for i, v in enumerate(vals, start=1):
                self._cell(ws, r, i, _safe_text(v), color=C_INPUT, fmt=(FMT_NUM if i in (3, 4, 5, 7) else None))
            r += 1
        last = max(r - 1, hdr + 1)
        self._table(ws, 'tblAircraft', hdr, last, 1, len(cols))
        self._named('Aircraft_id', 'Data', f'$A${hdr + 1}:$A${last}')
        self._named('Aircraft_battery', 'Data', f'$C${hdr + 1}:$C${last}')
        self._named('Aircraft_range', 'Data', f'$D${hdr + 1}:$D${last}')
        r += 2

        # --- CHARGERS (input) ---
        self._section(ws, r, 'CHARGERS', 'input — round-trippable')
        r += 1
        self._header_row(ws, r, ['Charger ID', 'Name', 'Power (kW)'])
        hdr = r
        r += 1
        for c in self.chargers:
            self._cell(ws, r, 1, _safe_text(c.get('id')), color=C_INPUT)
            self._cell(ws, r, 2, _safe_text(c.get('name')), color=C_INPUT)
            self._cell(ws, r, 3, _num(c.get('power_kw')), color=C_INPUT, fmt=FMT_KW)
            r += 1
        last = max(r - 1, hdr + 1)
        self._table(ws, 'tblChargers', hdr, last, 1, 3)
        self._named('Charger_id', 'Data', f'$A${hdr + 1}:$A${last}')
        self._named('Charger_power', 'Data', f'$C${hdr + 1}:$C${last}')
        r += 2

        # --- SETTINGS (input, SINGLE-ROW table: one column per setting) ---
        self._section(ws, r, 'SETTINGS', 'input — round-trippable')
        r += 1
        s = self.settings
        rp = s.get('routingPadding') or {}
        cols = [
            ('Charge target', _num(s.get('chargeTarget')), FMT_PCT, 'Settings_chargeTarget'),
            ('Tariff (EUR/kWh)', self.tariff, '€0.00', 'Settings_tariff'),
            ('Routing padding', 'on' if rp.get('enabled') else 'off', None, None),
            ('Padding factor', _num(rp.get('factor')) or 1.0, '0.00', None),
            ('SID/STAR (km/leg)', _num(s.get('sidStarPaddingKm')) or 0, FMT_NUM, None),
            ('Alternate reserve', 'on' if s.get('alternateReserve') else 'off', None, None),
            ('Grid factor', _num(s.get('gridDemandFactor')) or 1.0, '0.00', None),
            ('Day start', _clock(DAY_START_MIN), None, None),
            ('Day end', _clock(DAY_END_MIN), None, None),
            ('Realisation low', self.realisation_low, FMT_PCT, 'Settings_realisationLow'),
            ('Realisation high', self.realisation_high, FMT_PCT, 'Settings_realisationHigh'),
            ('Procurement (EUR/kWh)', self.procurement, '€0.00', 'Settings_procurement'),
        ]
        self._header_row(ws, r, [c[0] for c in cols])
        hdr = r
        r += 1
        from openpyxl.utils import get_column_letter
        for i, (_label, val, fmt, named) in enumerate(cols, start=1):
            self._cell(ws, r, i, val, color=C_INPUT, fmt=fmt)
            if named:
                self._named(named, 'Data', f'${get_column_letter(i)}${r}')
        self._table(ws, 'tblSettings', hdr, r, 1, len(cols))
        return ws

    # ---- Overview --------------------------------------------------------------
    def _overview(self):
        ws = self.wb.create_sheet('Overview')
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = TAB_NAVY
        self._widths(ws, [26, 15, 15, 18, 14, 3])
        self._cell(ws, 1, 1, 'Network overview', bold=True, size=18, color=NAVY)

        n_first, n_last = self._ap_first, self._ap_last
        daily_col = f"'Data'!$E${n_first}:$E${n_last}"
        peak_col = f"'Data'!$F${n_first}:$F${n_last}"

        # --- KPI cards: 4 across, navy fill, big white numbers ---
        cards = [
            ('Daily energy', f'=SUM({daily_col})', FMT_KWH),
            ('Peak demand', f'=MAX({peak_col})', FMT_KW),
            ('Energy / year', f'=SUM({daily_col})*365/1000', '#,##0 "MWh"'),
            ('Gross margin / yr', f'=SUM({daily_col})*365*(Settings_tariff-Settings_procurement)', FMT_EUR),
        ]
        from openpyxl.styles import PatternFill, Font, Alignment
        col = 1
        for label, formula, fmt in cards:
            for rr in (3, 4):       # fill both rows of the card area defensively
                for cc in (col,):
                    ws.cell(row=rr, column=cc).fill = PatternFill('solid', fgColor=NAVY)
            num = self._cell(ws, 3, col, formula, bold=True, size=15, color=WHITE, fill=NAVY,
                             fmt=fmt, align='center')
            lab = self._cell(ws, 4, col, label, size=9, color=WHITE, fill=NAVY, align='center')
            col += 1
        ws.row_dimensions[3].height = 26
        ws.row_dimensions[4].height = 14

        # --- scenario lines (live formulas) ---
        scen = [
            ('Gross revenue / year', f'=SUM({daily_col})*365*Settings_tariff', FMT_EUR),
            ('Revenue band (realisation)', f'=SUM({daily_col})*365*Settings_tariff*Settings_realisationLow',
             FMT_EUR),
            ('Energy cost / year', f'=SUM({daily_col})*365*Settings_procurement', FMT_EUR),
        ]
        r = 6
        for label, formula, fmt in scen:
            self._cell(ws, r, 1, label, bold=True, color=NAVY)
            self._cell(ws, r, 2, formula, color=C_FORMULA, fmt=fmt)
            if 'band' in label:
                self._cell(ws, r, 3, f'=SUM({daily_col})*365*Settings_tariff*Settings_realisationHigh',
                           color=C_FORMULA, fmt=FMT_EUR)
                self._cell(ws, r, 4, 'low – high', size=9, color=C_MUTED)
            r += 1

        # --- hyperlinked airport index (busiest first, mirrors the Data table) ---
        r += 1
        self._cell(ws, r, 1, 'AIRPORTS', bold=True, color=NAVY)
        self._cell(ws, r, 2, 'click a name to open its tab', italic=True, size=9, color=C_MUTED)
        r += 1
        self._header_row(ws, r, ['Airport', 'ICAO', 'Daily (kWh)', 'Peak (kW)', 'Revenue/day'])
        r += 1
        for ref, data_row in zip(self._airport_refs, self._ap_rows):
            self._hyperlink(ws, r, 1, ref['name'], ref['sheet'])
            self._cell(ws, r, 2, _safe_text(ref['ident']), color=C_MUTED)
            self._cell(ws, r, 3, f"='Data'!E{data_row}", color=C_LINK, fmt=FMT_NUM)
            self._cell(ws, r, 4, f"='Data'!F{data_row}", color=C_LINK, fmt=FMT_NUM)
            self._cell(ws, r, 5, f"='Data'!H{data_row}", color=C_LINK, fmt=FMT_EUR)
            r += 1

        # --- brand bar charts (sorted desc already; horizontal reads better) ---
        n = len(self._airport_refs)
        if n >= 2:
            self._add_airport_bar(ws, 'Daily energy per airport (kWh)', data_col=5, anchor='G3', n=n)
            self._add_airport_bar(ws, 'Peak demand per airport (kW)', data_col=6,
                                  anchor=f'G{6 + max(12, int(n * 1.1))}', n=n)
        return ws

    # ---- charts -------------------------------------------------------------
    def _add_donut(self, ws, first, last, anchor):
        if not first or not last or last < first:
            return
        from openpyxl.chart import DoughnutChart, Reference
        from openpyxl.chart.series import DataPoint
        from openpyxl.chart.shapes import GraphicalProperties
        ch = DoughnutChart()
        ch.title = 'Energy by aircraft type'
        ch.height, ch.width = 7.5, 12
        ch.holeSize = 55
        ch.varyColors = False
        data = Reference(ws, min_col=2, min_row=first - 1, max_row=last)
        cats = Reference(ws, min_col=1, min_row=first, max_row=last)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        # house palette per slice (instead of the rainbow default)
        pts = []
        for i in range(last - first + 1):
            dp = DataPoint(idx=i)
            dp.graphicalProperties = GraphicalProperties(solidFill=PALETTE[i % len(PALETTE)])
            pts.append(dp)
        ch.series[0].data_points = pts
        ws.add_chart(ch, anchor)

    def _add_step_curve(self, ws, first, last, anchor):
        """The load profile as a scatter-with-straight-lines over Excel-time x —
        a true step on a time-proportional axis (a category line chart would
        space unequal intervals equally and draw diagonals)."""
        from openpyxl.chart import ScatterChart, Series, Reference
        from openpyxl.chart.marker import Marker
        ch = ScatterChart()
        ch.title = 'Daily load profile (kW)'
        ch.height, ch.width = 8, 15
        ch.scatterStyle = 'lineMarker'
        ch.x_axis.number_format = FMT_TIME
        ch.x_axis.delete = False
        ch.y_axis.delete = False
        ch.y_axis.title = 'kW'
        ch.legend = None
        xref = Reference(ws, min_col=1, min_row=first, max_row=last)
        yref = Reference(ws, min_col=2, min_row=first, max_row=last)
        s = Series(yref, xref, title='kW')
        s.marker = Marker(symbol='none')
        s.smooth = False
        self._solid_series(s, PALETTE[0], line=True)
        ch.series.append(s)
        ws.add_chart(ch, anchor)

    def _add_airport_bar(self, ws, title, data_col, anchor, n):
        """Horizontal brand-blue bars over the Data sheet's Airports table."""
        from openpyxl.chart import BarChart, Reference
        dws = self.wb['Data']
        ch = BarChart()
        ch.type = 'bar'             # horizontal — 16 long airport names read sideways
        ch.title = title
        ch.height = max(7, 1.2 + 0.55 * n)
        ch.width = 15
        ch.legend = None
        ch.varyColors = False
        ch.gapWidth = 45
        data = Reference(dws, min_col=data_col, min_row=self._ap_first - 1, max_row=self._ap_last)
        cats = Reference(dws, min_col=1, min_row=self._ap_first, max_row=self._ap_last)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        self._solid_series(ch.series[0], PALETTE[0])
        ws.add_chart(ch, anchor)

    # ---- assembly -----------------------------------------------------------
    def build(self):
        self.wb.remove(self.wb.active)   # drop the default sheet
        for a in self.airports:          # busiest first (sorted in __init__)
            self._airport_detail(a)
        self._data_sheet()
        self._overview()
        # presentation order: Overview, Data, then the airport tabs
        order = ['Overview', 'Data'] + [r['sheet'] for r in self._airport_refs]
        self.wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 999)
        self.wb.active = 0
        buf = io.BytesIO()
        self.wb.save(buf)
        return buf.getvalue()


# ---- small value coercers ---------------------------------------------------
def _clock(minutes):
    try:
        m = max(0, int(round(float(minutes)))) % (24 * 60)
        return f'{m // 60:02d}:{m % 60:02d}'
    except (TypeError, ValueError):
        return ''


def _num(v):
    try:
        if v is None or v == '':
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


# Spreadsheet (CSV/Excel) formula injection: a user-supplied string that starts
# with a formula trigger becomes a live formula / DDE call when the workbook is
# opened or re-exported to CSV. Neutralise by prefixing a single apostrophe so
# the app treats the value as plain text. Only applied to USER-CONTROLLED values
# (names, idents, free text) — never to the code's own intended formula strings,
# so legitimate "=SUM(...)" cells keep working. Legitimate names never start with
# these characters, so the prefix is effectively invisible in normal use.
_FORMULA_LEAD = ('=', '+', '-', '@', '\t', '\r', '\n')


def _safe_text(v):
    if isinstance(v, str) and v[:1] in _FORMULA_LEAD:
        return "'" + v
    return v


def generate_xlsx(payload):
    """Render the workbook and return its bytes. Raises RuntimeError if openpyxl
    is unavailable (mirrors report.py's WeasyPrint handling)."""
    try:
        import openpyxl  # noqa: F401
    except ImportError as e:
        raise RuntimeError('openpyxl is not installed on the server. Install it with '
                           f'`pip install openpyxl`. Original error: {e}')
    return SpreadsheetBuilder(payload).build()
