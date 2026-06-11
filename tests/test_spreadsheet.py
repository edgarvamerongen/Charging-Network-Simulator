"""Unit tests for spreadsheet.py — build a small workbook from a representative
payload and assert the sheet structure, formulas and tab-name sanitisation.
Skipped cleanly if openpyxl isn't installed (it's a pinned requirement, so in a
provisioned env these always run)."""
import io
import os
import sys
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

from spreadsheet import generate_xlsx, SpreadsheetBuilder, _num, _clock, _safe_text


PAYLOAD = {
    'chargeRate': 0.55,
    'airports': [
        {
            'ident': 'EHLE', 'name': 'Lelystad Airport', 'lat': 52.46, 'lon': 5.52,
            'peakKw': 400,
            'chargers': [{'name': 'DC 400', 'power_kw': 400, 'count': 1}],
            'contribs': [
                {'role': 'home', 'other': 'Frankfurt', 'planeName': 'Beta Alia',
                 'tripType': 'retour', 'freqN': 2, 'freqUnit': 'day',
                 'energyPerFlight': 120, 'chargeMin': 30},
            ],
            'loadCurvePoints': [{'t': 420, 'kw': 0}, {'t': 480, 'kw': 400}, {'t': 540, 'kw': 0}],
        },
        # Hostile tab name: must be sanitised (no []:*?/\) and unique.
        {'ident': '', 'name': 'Bad[]:*?/\\Name', 'lat': None, 'lon': None,
         'peakKw': 0, 'chargers': [], 'contribs': []},
    ],
    'planes': [{'id': 'beta_plane', 'name': 'Beta Alia', 'battery_kwh': 260,
                'range_km': 463, 'speed_kmh': 270, 'seats': 5, 'load_kg': 500}],
    'chargers': [{'id': 'dc_400', 'name': 'DC 400', 'power_kw': 400}],
    'flightsFull': [{'id': 't1', 'planeId': 'beta_plane', 'planeName': 'Beta Alia',
                     'originIdent': 'EHLE', 'originName': 'Lelystad', 'originLat': 52.46,
                     'originLon': 5.52, 'destIdent': 'EDDF', 'destName': 'Frankfurt',
                     'destLat': 50.03, 'destLon': 8.57, 'stops': [], 'tripType': 'retour',
                     'multiLeg': False, 'chargerId': 'dc_400', 'chargerName': 'DC 400',
                     'freqN': 2, 'freqUnit': 'day'}],
    'modelSettings': {'chargeTarget': 0.8, 'routingPadding': {'enabled': True, 'factor': 1.1},
                      'sidStarPaddingKm': 15, 'alternateReserve': True, 'gridDemandFactor': 1.1},
}


@unittest.skipUnless(HAVE_OPENPYXL, 'openpyxl not installed')
class TestSpreadsheet(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.wb = openpyxl.load_workbook(io.BytesIO(generate_xlsx(PAYLOAD)))

    def test_sheet_structure(self):
        names = self.wb.sheetnames
        # Workbook v2: inputs + about collapse into the single Data sheet
        # (tables addressed by NAME there, not by sheet placement).
        for expected in ('Overview', 'Data', 'EHLE'):
            self.assertIn(expected, names)
        for gone in ('About', 'Flights', 'Aircraft', 'Chargers', 'Settings'):
            self.assertNotIn(gone, names)

    def test_hostile_tab_name_sanitised(self):
        for n in self.wb.sheetnames:
            self.assertFalse(any(ch in n for ch in '[]:*?/\\'), n)
        self.assertTrue(any(n.startswith('BadName') for n in self.wb.sheetnames),
                        self.wb.sheetnames)

    def test_flights_table_row(self):
        # v2: the flights table lives on Data, addressed by name (tblFlights).
        ws = self.wb['Data']
        tbl = ws.tables['tblFlights']
        first_col, first_row = openpyxl.utils.range_boundaries(tbl.ref)[:2]
        self.assertEqual(ws.cell(row=first_row, column=first_col).value, 'Flight ID')
        self.assertEqual(ws.cell(row=first_row + 1, column=first_col).value, 't1')
        self.assertEqual(ws.cell(row=first_row + 1, column=first_col + 2).value, 'Beta Alia')

    def test_airport_daily_total_is_formula(self):
        ws = self.wb['EHLE']
        formulas = [c.value for row in ws.iter_rows() for c in row
                    if isinstance(c.value, str) and c.value.startswith('=SUM(H')]
        self.assertTrue(formulas, 'expected a =SUM(H…) daily-total formula')

    def test_overview_revenue_uses_named_tariff(self):
        ws = self.wb['Overview']
        vals = [c.value for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and 'Settings_tariff' in c.value]
        self.assertTrue(vals, 'expected revenue formulas bound to Settings_tariff')

    def test_tariff_value_round_trips(self):
        # v2: settings are a single-row table on Data; the tariff cell carries
        # the defined name Settings_tariff — resolve it, don't pin placement.
        dn = self.wb.defined_names['Settings_tariff']
        (sheet, ref), = dn.destinations
        self.assertAlmostEqual(self.wb[sheet][ref].value, 0.55)

    def test_tab_name_collisions_get_suffix(self):
        b = SpreadsheetBuilder({})
        self.assertEqual(b._tab('EHAM'), 'EHAM')
        self.assertEqual(b._tab('EHAM'), 'EHAM (2)')
        self.assertEqual(b._tab('eham'), 'eham (3)')   # uniqueness is case-insensitive


@unittest.skipUnless(HAVE_OPENPYXL, 'openpyxl not installed')
class TestFormulaInjection(unittest.TestCase):
    """A user-named plane/charger/airport must never become a live formula in
    the exported workbook (spreadsheet/CSV injection)."""

    def test_malicious_names_are_neutralised_to_text(self):
        evil = '=1+1'
        ddE = '=cmd|\'/c calc\'!A0'
        payload = {
            'chargeRate': 0.6,
            'airports': [{
                'ident': 'EHAM', 'name': evil, 'lat': 52.3, 'lon': 4.76, 'peakKw': 100,
                'chargers': [{'name': ddE, 'power_kw': 100, 'count': 1}],
                'contribs': [{'role': 'origin', 'other': '=HYPERLINK("http://x")',
                              'planeName': '@SUM(A1)', 'tripType': 'one-way',
                              'freqN': 1, 'freqUnit': 'day', 'energyPerFlight': 10, 'chargeMin': 5}],
            }],
            'planes': [{'id': 'p', 'name': '+evil', 'battery_kwh': 100, 'range_km': 200, 'speed_kmh': 200}],
            'chargers': [{'id': 'c', 'name': ddE, 'power_kw': 100}],
            'flightsFull': [{'id': '-1+2', 'planeId': 'p', 'planeName': '@SUM(A1)',
                             'originIdent': 'EHAM', 'originName': evil, 'destIdent': 'EDDF',
                             'destName': 'Frankfurt', 'tripType': 'one-way',
                             'chargerId': 'c', 'chargerName': ddE, 'freqN': 1, 'freqUnit': 'day'}],
        }
        wb = openpyxl.load_workbook(io.BytesIO(generate_xlsx(payload)))
        live_formulas = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and isinstance(cell.value, str):
                        # an intended formula is fine; a user string leaking through is not
                        if any(tok in cell.value for tok in ('cmd|', 'HYPERLINK', '1+1')):
                            live_formulas.append((ws.title, cell.coordinate, cell.value))
        self.assertEqual(live_formulas, [], f'user input became a live formula: {live_formulas}')

    def test_safe_text_prefixes_only_dangerous_leads(self):
        self.assertEqual(_safe_text('=1+1'), "'=1+1")
        self.assertEqual(_safe_text('+1'), "'+1")
        self.assertEqual(_safe_text('-1'), "'-1")
        self.assertEqual(_safe_text('@x'), "'@x")
        self.assertEqual(_safe_text('Lelystad'), 'Lelystad')   # normal names untouched
        self.assertEqual(_safe_text(None), None)
        self.assertEqual(_safe_text(42), 42)


class TestCoercers(unittest.TestCase):
    def test_num(self):
        self.assertEqual(_num('3.5'), 3.5)
        self.assertIsNone(_num(None))
        self.assertIsNone(_num(''))
        self.assertIsNone(_num('abc'))

    def test_clock(self):
        self.assertEqual(_clock(480), '08:00')
        self.assertEqual(_clock(0), '00:00')
        self.assertEqual(_clock(24 * 60 + 30), '00:30')   # wraps
        self.assertEqual(_clock('nope'), '')


if __name__ == '__main__':
    unittest.main()
